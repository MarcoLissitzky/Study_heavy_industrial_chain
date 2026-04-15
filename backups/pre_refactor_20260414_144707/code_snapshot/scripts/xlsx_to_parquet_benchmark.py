import argparse
import json
import math
import re
import time
import traceback
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path

import polars as pl
from openpyxl import load_workbook


@dataclass(frozen=True)
class XlsxMeta:
    path: str
    province: str
    city: str
    shard: int
    bytes: int


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_xlsx_meta(path: Path) -> XlsxMeta:
    stem = path.stem
    parts = [p.strip() for p in stem.split("_") if p.strip() != ""]
    if len(parts) < 2:
        raise ValueError(f"Unrecognized xlsx filename (expected 至少 省_市.xlsx): {path.name}")
    shard = 1
    if len(parts) >= 3 and parts[-1].isdigit():
        shard = int(parts[-1])
        province = parts[0]
        city = "_".join(parts[1:-1]) or "-"
    else:
        province = parts[0]
        city = "_".join(parts[1:]) or "-"
    return XlsxMeta(path=str(path), province=province, city=city, shard=shard, bytes=path.stat().st_size)


def iter_xlsx_files(input_dir: Path) -> list[XlsxMeta]:
    metas: list[XlsxMeta] = []
    for p in sorted(input_dir.glob("*.xlsx")):
        if p.name.startswith("~"):
            continue
        metas.append(parse_xlsx_meta(p))
    return metas


def pick_representative(metas: list[XlsxMeta], n: int, *, pick: str = "mixed") -> list[XlsxMeta]:
    if n <= 0:
        return []
    if len(metas) <= n:
        return metas
    uniq: dict[str, XlsxMeta] = {m.path: m for m in metas}
    if pick == "largest":
        metas_sorted = sorted(uniq.values(), key=lambda x: x.bytes, reverse=True)
    elif pick == "smallest":
        metas_sorted = sorted(uniq.values(), key=lambda x: x.bytes)
    else:
        metas_sorted = sorted(uniq.values(), key=lambda x: x.bytes, reverse=True)

    want_large = max(1, min(5, n // 2 + 1))
    selected: list[XlsxMeta] = []
    seen_pc: set[tuple[str, str]] = set()
    for m in metas_sorted:
        if len(selected) >= want_large:
            break
        key = (m.province, m.city)
        if key in seen_pc:
            continue
        selected.append(m)
        seen_pc.add(key)
    remaining = [m for m in metas_sorted if m not in selected]
    for q in [0.10, 0.30, 0.50, 0.70, 0.90]:
        if len(selected) >= n or not remaining:
            break
        idx = min(len(remaining) - 1, max(0, int(math.floor(q * (len(remaining) - 1)))))
        cand = remaining[idx]
        key = (cand.province, cand.city)
        if key not in seen_pc:
            selected.append(cand)
            seen_pc.add(key)
    for m in remaining:
        if len(selected) >= n:
            break
        if m not in selected:
            selected.append(m)
    out: list[XlsxMeta] = []
    seen_paths: set[str] = set()
    for m in selected:
        if m.path in seen_paths:
            continue
        seen_paths.add(m.path)
        out.append(m)
        if len(out) >= n:
            break
    return out


def normalize_columns(cols: list[object]) -> list[str]:
    out: list[str] = []
    seen: dict[str, int] = {}
    for c in cols:
        s = "" if c is None else str(c)
        s = s.replace("\u3000", " ").strip()
        s = re.sub(r"\s+", " ", s)
        if not s:
            s = "unnamed"
        k = seen.get(s, 0)
        seen[s] = k + 1
        if k:
            s = f"{s}__{k+1}"
        out.append(s)
    return out


def cell_to_str(v: object) -> object:
    if v is None:
        return None
    try:
        iso = getattr(v, "isoformat", None)
        if callable(iso):
            return iso()
    except Exception:
        pass
    s = str(v).replace("\u3000", " ").strip()
    return s if s else None


def _append_jsonl(path: Path, obj: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def load_latest_manifest_status(manifest_path: Path) -> dict[str, dict[str, object]]:
    if not manifest_path.exists():
        return {}
    latest: dict[str, dict[str, object]] = {}
    with manifest_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue
            sf = rec.get("source_file")
            if isinstance(sf, str):
                latest[sf] = rec
    return latest


def write_summary(summary_path: Path, data: dict[str, object]) -> None:
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = summary_path.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(summary_path)


def stream_xlsx_to_parquet_parts(meta: XlsxMeta, out_root: str, chunk_rows: int, zstd_level: int, sheet_name: str | None) -> dict[str, object]:
    t0 = time.perf_counter()
    path = Path(meta.path)
    out_dir = Path(out_root) / f"province={meta.province}" / f"city={meta.city}"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        header_raw = next(rows, None)
        if header_raw is None:
            return {
                "source_file": meta.path,
                "province": meta.province,
                "city": meta.city,
                "shard": meta.shard,
                "xlsx_bytes": meta.bytes,
                "rows": 0,
                "parquet_parts": 0,
                "parquet_bytes": 0,
                "seconds": time.perf_counter() - t0,
                "note": "empty_sheet",
            }
        cols = normalize_columns(list(header_raw))
        buffer: list[tuple[object, ...]] = []
        total_rows = 0
        parquet_bytes = 0
        part_idx = 0

        def flush() -> None:
            nonlocal part_idx, parquet_bytes, total_rows
            if not buffer:
                return
            col_arrays = list(zip(*buffer, strict=False))
            data: dict[str, list[object]] = {}
            for name, arr in zip(cols, col_arrays, strict=False):
                data[name] = [cell_to_str(x) for x in arr]
            df = pl.DataFrame(data, schema={k: pl.Utf8 for k in data.keys()})
            df = df.with_columns(
                pl.lit(meta.province).alias("province"),
                pl.lit(meta.city).alias("city"),
                pl.lit(meta.shard).alias("shard"),
                pl.lit(path.name).alias("source_file"),
            )
            part_path = out_dir / f"part-{meta.shard:04d}-{part_idx:05d}.parquet"
            tmp_path = out_dir / f".part-{meta.shard:04d}-{part_idx:05d}.tmp"
            df.write_parquet(tmp_path, compression="zstd", compression_level=zstd_level, statistics=True, use_pyarrow=False)
            tmp_path.replace(part_path)
            parquet_bytes += part_path.stat().st_size
            total_rows += df.height
            part_idx += 1
            buffer.clear()

        for r in rows:
            if r is None:
                continue
            buffer.append(tuple(r))
            if len(buffer) >= chunk_rows:
                flush()
        flush()

        return {
            "source_file": meta.path,
            "province": meta.province,
            "city": meta.city,
            "shard": meta.shard,
            "xlsx_bytes": meta.bytes,
            "rows": int(total_rows),
            "parquet_parts": int(part_idx),
            "parquet_bytes": int(parquet_bytes),
            "seconds": float(time.perf_counter() - t0),
            "sheet": ws.title,
            "chunk_rows": int(chunk_rows),
            "zstd_level": int(zstd_level),
        }
    finally:
        wb.close()


def main() -> int:
    p = argparse.ArgumentParser(description="Streaming XLSX -> partitioned Parquet with resume/progress")
    p.add_argument("--input-dir", required=True)
    p.add_argument("--out-dir", required=True)
    p.add_argument("--n", type=int, default=6)
    p.add_argument("--all", action="store_true", help="Process all files instead of representative subset")
    p.add_argument("--pick", choices=["mixed", "largest", "smallest"], default="mixed")
    p.add_argument("--min-bytes", type=int, default=0)
    p.add_argument("--max-bytes", type=int, default=0)
    p.add_argument("--chunk-rows", type=int, default=50_000)
    p.add_argument("--zstd-level", type=int, default=6)
    p.add_argument("--sheet", default="")
    p.add_argument("--list-only", action="store_true")
    p.add_argument("--workers", type=int, default=1)
    p.add_argument("--retry-failed", action="store_true", help="Re-run files that previously failed")
    args = p.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    run_dir = out_dir / "run"
    manifest_path = run_dir / "manifest.jsonl"
    summary_path = run_dir / "summary.json"
    results_path = out_dir / "benchmark_results.jsonl"
    out_dir.mkdir(parents=True, exist_ok=True)

    metas = iter_xlsx_files(input_dir)
    if not metas:
        raise SystemExit(f"No .xlsx files found in: {input_dir}")
    if args.min_bytes or args.max_bytes:
        max_b = args.max_bytes if args.max_bytes > 0 else None
        metas = [m for m in metas if m.bytes >= args.min_bytes and (max_b is None or m.bytes <= max_b)]
        if not metas:
            raise SystemExit("No .xlsx files after --min-bytes/--max-bytes filters.")

    selected = metas if args.all else pick_representative(metas, max(1, args.n), pick=args.pick)
    latest = load_latest_manifest_status(manifest_path)
    to_process: list[XlsxMeta] = []
    for m in selected:
        prev = latest.get(m.path, {})
        st = prev.get("status")
        if st == "done":
            continue
        if st == "failed" and not args.retry_failed:
            continue
        to_process.append(m)

    if args.list_only:
        for m in to_process:
            print(f"{m.bytes}\t{m.province}\t{m.city}\t{Path(m.path).name}", flush=True)
        return 0

    total_files = len(to_process)
    if total_files == 0:
        print("No files need processing (all done, or failed and --retry-failed not enabled).")
        return 0

    sheet_name = args.sheet.strip() or None
    total_t0 = time.perf_counter()
    completed = 0
    failed = 0
    total_rows = 0
    total_parquet_bytes = 0
    total_xlsx_bytes = sum(x.bytes for x in to_process)

    for m in to_process:
        _append_jsonl(manifest_path, {"ts": now_iso(), "status": "pending", "source_file": m.path, **asdict(m)})

    with ProcessPoolExecutor(max_workers=max(1, args.workers)) as ex:
        future_map = {}
        for m in to_process:
            _append_jsonl(manifest_path, {"ts": now_iso(), "status": "running", "source_file": m.path, **asdict(m)})
            fut = ex.submit(stream_xlsx_to_parquet_parts, m, str(out_dir), int(args.chunk_rows), int(args.zstd_level), sheet_name)
            future_map[fut] = m

        for fut in as_completed(future_map):
            m = future_map[fut]
            try:
                res = fut.result()
                completed += 1
                total_rows += int(res["rows"])
                total_parquet_bytes += int(res["parquet_bytes"])
                _append_jsonl(results_path, res)
                _append_jsonl(manifest_path, {"ts": now_iso(), "status": "done", **res})
                print(
                    f"done\t{Path(m.path).name}\t{completed}/{total_files}\trows={res['rows']}\t"
                    f"sec={res['seconds']:.1f}\tparquet_gb={res['parquet_bytes']/1e9:.3f}",
                    flush=True,
                )
            except Exception as e:
                failed += 1
                err = {"type": type(e).__name__, "message": str(e), "traceback": traceback.format_exc(limit=8)}
                _append_jsonl(manifest_path, {"ts": now_iso(), "status": "failed", "source_file": m.path, "error": err, **asdict(m)})
                print(f"failed\t{Path(m.path).name}\t{failed}\t{type(e).__name__}: {e}", flush=True)

            elapsed = time.perf_counter() - total_t0
            done_all = completed + failed
            rate_files = done_all / elapsed if elapsed > 0 else 0.0
            eta_sec = (total_files - done_all) / rate_files if rate_files > 0 else None
            summary = {
                "ts": now_iso(),
                "input_dir": str(input_dir),
                "out_dir": str(out_dir),
                "total_files": total_files,
                "completed": completed,
                "failed": failed,
                "remaining": total_files - done_all,
                "elapsed_sec": elapsed,
                "eta_sec": eta_sec,
                "rows": total_rows,
                "xlsx_bytes": total_xlsx_bytes,
                "parquet_bytes": total_parquet_bytes,
                "compression_ratio": (total_parquet_bytes / total_xlsx_bytes) if total_xlsx_bytes else None,
                "workers": max(1, args.workers),
            }
            write_summary(summary_path, summary)
            eta_txt = f"{eta_sec/60:.1f}m" if eta_sec is not None else "NA"
            print(
                f"progress\t{done_all}/{total_files}\tcompleted={completed}\tfailed={failed}\t"
                f"rows={total_rows}\teta={eta_txt}",
                flush=True,
            )

    elapsed = time.perf_counter() - total_t0
    print(f"manifest\t{manifest_path}")
    print(f"summary\t{summary_path}")
    print(f"results\t{results_path}")
    print(f"total_seconds\t{elapsed:.1f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
